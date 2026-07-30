import csv
import io
from datetime import datetime, timezone

from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy.orm import Session

from app.crud import rubric as rubric_crud
from app.services import analytics_service, scoring_service


def _gather_report_data(db: Session) -> dict:
    rankings, _ = analytics_service.get_leaderboard(db, page=1, page_size=100000)
    judge_stats = scoring_service.compute_judge_stats(db)
    criterion_averages = analytics_service.get_criterion_averages(db)
    rubric = rubric_crud.get_active(db)
    criteria_names = [c.name for c in rubric.criteria] if rubric else []
    return {
        "rankings": rankings,
        "judge_stats": judge_stats,
        "criterion_averages": criterion_averages,
        "criteria_names": criteria_names,
        "generated_at": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC"),
    }


def build_csv(db: Session) -> bytes:
    data = _gather_report_data(db)
    buffer = io.StringIO()
    writer = csv.writer(buffer)

    writer.writerow([f"Hackathon Evaluation Report - generated {data['generated_at']}"])
    writer.writerow([])

    writer.writerow(["FINAL RANKINGS"])
    writer.writerow(["Rank", "Project", "Problem Statement", "Overall Score", "Std Dev", "Reviews Completed", "Flagged"])
    for entry in data["rankings"]:
        writer.writerow(
            [
                entry.rank,
                entry.project_title,
                entry.problem_statement or "",
                entry.overall_score if entry.overall_score is not None else "",
                entry.std_dev if entry.std_dev is not None else "",
                entry.reviews_completed,
                "Yes" if entry.is_flagged else "No",
            ]
        )
    writer.writerow([])

    writer.writerow(["SUBMISSION CRITERION BREAKDOWN"])
    writer.writerow(["Project"] + data["criteria_names"])
    for entry in data["rankings"]:
        writer.writerow([entry.project_title] + [entry.criterion_scores.get(name, "") for name in data["criteria_names"]])
    writer.writerow([])

    writer.writerow(["JUDGE STATISTICS"])
    writer.writerow(
        ["Judge", "Assigned", "Completed", "Pending", "Avg Score Given", "Std Dev", "Avg Review Time (s)", "Harsh", "Lenient", "High Variance"]
    )
    for js in data["judge_stats"]:
        writer.writerow(
            [
                js.judge.full_name or js.judge.username,
                js.reviews_assigned,
                js.reviews_completed,
                js.reviews_pending,
                js.average_score_given if js.average_score_given is not None else "",
                js.std_dev_given if js.std_dev_given is not None else "",
                js.average_review_time_seconds if js.average_review_time_seconds is not None else "",
                "Yes" if js.is_harsh else "No",
                "Yes" if js.is_lenient else "No",
                "Yes" if js.is_high_variance else "No",
            ]
        )
    writer.writerow([])

    writer.writerow(["CRITERION AVERAGES"])
    writer.writerow(["Criterion", "Average Score"])
    for c in data["criterion_averages"]:
        writer.writerow([c.criterion_name, c.average_score])

    return buffer.getvalue().encode("utf-8-sig")


def build_xlsx(db: Session) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    data = _gather_report_data(db)
    wb = Workbook()
    bold = Font(bold=True)

    ws = wb.active
    ws.title = "Rankings"
    ws.append(["Rank", "Project", "Problem Statement", "Overall Score", "Std Dev", "Reviews Completed", "Flagged"])
    for cell in ws[1]:
        cell.font = bold
    for entry in data["rankings"]:
        ws.append(
            [
                entry.rank,
                entry.project_title,
                entry.problem_statement or "",
                entry.overall_score,
                entry.std_dev,
                entry.reviews_completed,
                "Yes" if entry.is_flagged else "No",
            ]
        )

    ws2 = wb.create_sheet("Submission Statistics")
    ws2.append(["Project"] + data["criteria_names"])
    for cell in ws2[1]:
        cell.font = bold
    for entry in data["rankings"]:
        ws2.append([entry.project_title] + [entry.criterion_scores.get(name) for name in data["criteria_names"]])

    ws3 = wb.create_sheet("Judge Statistics")
    ws3.append(
        ["Judge", "Assigned", "Completed", "Pending", "Avg Score Given", "Std Dev", "Avg Review Time (s)", "Harsh", "Lenient", "High Variance"]
    )
    for cell in ws3[1]:
        cell.font = bold
    for js in data["judge_stats"]:
        ws3.append(
            [
                js.judge.full_name or js.judge.username,
                js.reviews_assigned,
                js.reviews_completed,
                js.reviews_pending,
                js.average_score_given,
                js.std_dev_given,
                js.average_review_time_seconds,
                "Yes" if js.is_harsh else "No",
                "Yes" if js.is_lenient else "No",
                "Yes" if js.is_high_variance else "No",
            ]
        )

    ws4 = wb.create_sheet("Criterion Breakdown")
    ws4.append(["Criterion", "Average Score"])
    for cell in ws4[1]:
        cell.font = bold
    for c in data["criterion_averages"]:
        ws4.append([c.criterion_name, c.average_score])

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def build_pdf(db: Session) -> bytes:
    data = _gather_report_data(db)
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
    styles = getSampleStyleSheet()
    story = [Paragraph("Hackathon Evaluation Report", styles["Title"]), Paragraph(data["generated_at"], styles["Normal"]), Spacer(1, 16)]

    def add_table(title: str, header: list[str], rows: list[list]):
        story.append(Paragraph(title, styles["Heading2"]))
        table_data = [header] + rows
        table = Table(table_data, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f3f4f6")]),
                ]
            )
        )
        story.append(table)
        story.append(Spacer(1, 16))

    add_table(
        "Final Rankings",
        ["Rank", "Project", "Problem Statement", "Overall", "Std Dev", "Reviews", "Flagged"],
        [
            [
                e.rank,
                e.project_title,
                e.problem_statement or "",
                f"{e.overall_score:.2f}" if e.overall_score is not None else "-",
                f"{e.std_dev:.2f}" if e.std_dev is not None else "-",
                e.reviews_completed,
                "Yes" if e.is_flagged else "No",
            ]
            for e in data["rankings"]
        ],
    )

    add_table(
        "Judge Statistics",
        ["Judge", "Assigned", "Completed", "Avg Score", "Std Dev", "Flags"],
        [
            [
                js.judge.full_name or js.judge.username,
                js.reviews_assigned,
                js.reviews_completed,
                f"{js.average_score_given:.2f}" if js.average_score_given is not None else "-",
                f"{js.std_dev_given:.2f}" if js.std_dev_given is not None else "-",
                ", ".join(
                    filter(None, [js.is_harsh and "Harsh", js.is_lenient and "Lenient", js.is_high_variance and "High Variance"])
                )
                or "-",
            ]
            for js in data["judge_stats"]
        ],
    )

    add_table(
        "Criterion Breakdown",
        ["Criterion", "Average Score"],
        [[c.criterion_name, f"{c.average_score:.2f}"] for c in data["criterion_averages"]],
    )

    doc.build(story)
    return buffer.getvalue()
